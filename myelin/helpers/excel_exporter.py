"""
Excel Exporter for MyelinOutput

This module provides functionality to export MyelinOutput data to Excel files,
making it easier to share results with business users for analysis.

The exporter handles:
- Nested Pydantic models by flattening them with prefixed column names
- Lists/arrays by creating separate sub-tables within worksheets
- Multiple output types (IOCE, MCE, MSDRG, pricers) as separate worksheets
- A summary sheet with key metrics from all modules
"""

from datetime import datetime
from enum import Enum
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING, Any

from pydantic import BaseModel

if TYPE_CHECKING:
    from myelin.core import MyelinOutput
    from myelin.input.claim import Claim

try:
    import openpyxl  # type: ignore[import-not-found]
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore[import-not-found]
    from openpyxl.utils import get_column_letter  # type: ignore[import-not-found]
    from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-not-found]

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _ensure_openpyxl() -> None:
    """Ensure openpyxl is available, raise ImportError with helpful message if not."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel export functionality. "
            "Install it with: pip install openpyxl"
        )


# Style constants
HEADER_FILL = None
HEADER_FONT = None
HEADER_ALIGNMENT = None
THIN_BORDER = None
WRAP_ALIGNMENT = None


def _init_styles() -> None:
    """Initialize styles after ensuring openpyxl is available."""
    global HEADER_FILL, HEADER_FONT, HEADER_ALIGNMENT, THIN_BORDER, WRAP_ALIGNMENT
    if HEADER_FILL is None:
        HEADER_FILL = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        HEADER_FONT = Font(bold=True, color="FFFFFF")
        HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
        THIN_BORDER = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def _format_value(value: Any) -> Any:
    """Format a value for Excel export."""
    if value is None:
        return ""
    if isinstance(value, Enum):
        return value.value
    if isinstance(value, datetime):
        return value
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, (list, dict)):
        # For nested structures that weren't flattened, convert to string
        return str(value) if value else ""
    return value


def _is_edit_list(field_name: str, field_value: list) -> bool:
    """Check if a field is an edit list that should be concatenated."""
    # Check if the field name suggests it's an edit list
    if "edit_list" in field_name.lower():
        return True
    # Also check if items have 'edit' attribute (IoceOutputEdit pattern)
    if field_value and hasattr(field_value[0], "edit"):
        return True
    return False


def _concatenate_edit_list(items: list) -> str:
    """
    Concatenate edit list items into a newline-delimited string.

    Handles IoceOutputEdit objects which have 'edit' and 'description' fields.
    """
    if not items:
        return ""

    parts = []
    for item in items:
        if hasattr(item, "edit") and hasattr(item, "description"):
            # IoceOutputEdit format
            edit_str = str(item.edit) if item.edit else ""
            desc_str = str(item.description) if item.description else ""
            if edit_str and desc_str:
                parts.append(f"{edit_str}: {desc_str}")
            elif edit_str:
                parts.append(edit_str)
            elif desc_str:
                parts.append(desc_str)
        elif hasattr(item, "edit"):
            edit_str = str(item.edit) if item.edit else ""
            if edit_str:
                parts.append(edit_str)
        else:
            # Fallback: just convert to string
            parts.append(str(item))

    return "\n".join(parts) if parts else ""


def _is_simple_string_list(field_name: str, field_value: list) -> bool:
    """
    Check if a field is a simple list of strings that should be joined.

    This handles fields like 'modifiers', 'cond_codes', 'demo_codes', etc.
    """
    if not field_value:
        return False
    # Check if all items are strings (or string-like primitives)
    if all(isinstance(item, str) for item in field_value):
        return True
    return False


def _concatenate_string_list(items: list[str], delimiter: str = "; ") -> str:
    """
    Concatenate a list of strings into a single delimited string.

    Args:
        items: List of strings to join
        delimiter: Delimiter to use between items (default: "; ")

    Returns:
        Joined string
    """
    if not items:
        return ""
    # Filter out empty strings and join
    non_empty = [str(item) for item in items if item]
    return delimiter.join(non_empty) if non_empty else ""


def _flatten_model(
    model: BaseModel, prefix: str = "", max_depth: int = 4
) -> dict[str, Any]:
    """
    Flatten a Pydantic model into a dictionary with prefixed keys.

    Args:
        model: The Pydantic model to flatten
        prefix: Prefix to add to keys (used for nested models)
        max_depth: Maximum recursion depth for nested models

    Returns:
        Flattened dictionary with prefixed keys
    """
    result: dict[str, Any] = {}

    if max_depth <= 0:
        return {prefix.rstrip("_"): str(model)}

    for field_name, field_value in model:
        key = f"{prefix}{field_name}" if prefix else field_name

        if field_value is None:
            result[key] = ""
        elif isinstance(field_value, BaseModel):
            # Recursively flatten nested models
            nested = _flatten_model(field_value, f"{key}_", max_depth - 1)
            result.update(nested)
        elif isinstance(field_value, list):
            # Check if this is an edit list that should be concatenated
            if field_value and _is_edit_list(field_name, field_value):
                result[key] = _concatenate_edit_list(field_value)
            # Check if this is a simple string list (like modifiers, cond_codes)
            elif field_value and _is_simple_string_list(field_name, field_value):
                result[key] = _concatenate_string_list(field_value)
            else:
                # Other lists will be handled separately, just note the count
                result[f"{key}_count"] = len(field_value)
        elif isinstance(field_value, dict):
            # Flatten dict with prefixed keys
            for dict_key, dict_value in field_value.items():
                result[f"{key}_{dict_key}"] = _format_value(dict_value)
        else:
            result[key] = _format_value(field_value)

    return result


def _extract_list_items(model: BaseModel) -> dict[str, list[BaseModel]]:
    """
    Extract all list fields from a model that contain BaseModel items.

    Returns:
        Dictionary mapping field names to lists of items
    """
    lists: dict[str, list[BaseModel]] = {}

    for field_name, field_value in model:
        if isinstance(field_value, list) and field_value:
            # Skip edit lists - they're handled inline via concatenation
            if _is_edit_list(field_name, field_value):
                continue
            # Skip simple string lists - they're handled inline via concatenation
            if _is_simple_string_list(field_name, field_value):
                continue
            # Check if items are BaseModels
            if isinstance(field_value[0], BaseModel):
                lists[field_name] = field_value

    return lists


def _style_header_row(ws: "Worksheet", row: int, col_count: int) -> None:
    """Apply header styling to a row."""
    _init_styles()
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _auto_adjust_column_widths(ws: "Worksheet") -> None:
    """Auto-adjust column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column_cells:
            if column_letter is None:
                column_letter = get_column_letter(cell.column)
            try:
                cell_value = str(cell.value) if cell.value else ""
                # Limit max width
                cell_length = min(len(cell_value), 50)
                max_length = max(max_length, cell_length)
            except (TypeError, AttributeError):
                pass
        if column_letter:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


def format_cell(key: str, value: Any, cell: "openpyxl.cell.cell.Cell") -> None:
    if isinstance(value, float):
        if "ratio" in key.lower() or "percent" in key.lower():
            # Percentage format
            cell.number_format = "0.00%"
        elif (
            "payment" in key.lower() or "amount" in key.lower() or "cost" in key.lower()
        ):
            # Currency format
            cell.number_format = '"$"#,##0.00'
        else:
            # We format to 4 decimal places but do not round the value itself
            # This allows users to keep full precesion but see a cleaner format
            cell.number_format = "#,##0.0000"
    elif isinstance(value, str):
        if "\n" in value or len(value) > 50:
            cell.alignment = WRAP_ALIGNMENT


def _write_model_to_sheet(
    ws: "Worksheet",
    model: BaseModel,
    title: str,
    start_row: int = 1,
) -> int:
    """
    Write a Pydantic model to a worksheet.

    Args:
        ws: The worksheet to write to
        model: The model to write
        title: Title for this section
        start_row: Row to start writing at

    Returns:
        The next available row after writing
    """
    _init_styles()
    current_row = start_row

    # Write section title
    ws.cell(row=current_row, column=1, value=title)
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
    current_row += 1

    # Flatten the model and write key-value pairs
    flat_data = _flatten_model(model)

    if flat_data:
        # Write headers
        ws.cell(row=current_row, column=1, value="Field")
        ws.cell(row=current_row, column=2, value="Value")
        _style_header_row(ws, current_row, 2)
        current_row += 1

        # Write data
        for key, value in flat_data.items():
            try:
                ws.cell(row=current_row, column=1, value=_humanize_key(key))
                cell = ws.cell(row=current_row, column=2, value=value)
                format_cell(key, value, cell)
                cell.border = THIN_BORDER
                ws.cell(row=current_row, column=1).border = THIN_BORDER
                current_row += 1
            except ValueError:
                # Skip values that cannot be written to Excel
                continue

    current_row += 1  # Add spacing

    # Handle list fields - create sub-tables
    list_fields = _extract_list_items(model)
    for field_name, items in list_fields.items():
        current_row = _write_list_to_sheet(ws, items, field_name, current_row)
        current_row += 1

    return current_row


def _write_list_to_sheet(
    ws: "Worksheet",
    items: list[BaseModel],
    title: str,
    start_row: int,
) -> int:
    """
    Write a list of Pydantic models as a table.

    Args:
        ws: The worksheet to write to
        items: List of models to write
        title: Title for this table
        start_row: Row to start writing at

    Returns:
        The next available row after writing
    """
    _init_styles()
    if not items:
        return start_row

    current_row = start_row

    # Write table title
    ws.cell(row=current_row, column=1, value=_humanize_key(title))
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    current_row += 1

    # Get all unique keys from all items
    all_keys: list[str] = []
    for item in items:
        flat = _flatten_model(item)
        for key in flat.keys():
            if key not in all_keys:
                all_keys.append(key)

    # Write headers
    for col, key in enumerate(all_keys, start=1):
        ws.cell(row=current_row, column=col, value=_humanize_key(key))
    _style_header_row(ws, current_row, len(all_keys))
    current_row += 1

    # Write data rows
    for item in items:
        flat = _flatten_model(item)
        for col, key in enumerate(all_keys, start=1):
            value = flat.get(key, "")
            cell = ws.cell(row=current_row, column=col, value=value)
            format_cell(key, value, cell)
            cell.border = THIN_BORDER
        current_row += 1

    return current_row


def _humanize_key(key: str) -> str:
    """Convert a snake_case key to a human-readable title."""
    # Replace underscores with spaces and title case
    return key.replace("_", " ").title()


def _create_summary_sheet(
    wb: "openpyxl.Workbook",
    output: "MyelinOutput",
    claim: "Claim | None" = None,
) -> None:
    """Create a summary sheet with key information from all modules."""
    _init_styles()
    ws = wb.active
    ws.title = "Summary"

    current_row = 1

    # Title
    ws.cell(row=current_row, column=1, value="Myelin Output Summary")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=16)
    current_row += 2

    # Claim summary if provided
    if claim is not None:
        ws.cell(row=current_row, column=1, value="Claim Information")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 1

        claim_info = [
            ("Claim ID", claim.claimid),
            (
                "From Date",
                claim.from_date.strftime("%Y-%m-%d") if claim.from_date else "",
            ),
            (
                "Thru Date",
                claim.thru_date.strftime("%Y-%m-%d") if claim.thru_date else "",
            ),
            ("Bill Type", claim.bill_type),
            (
                "Total Charges",
                f"${claim.total_charges:,.2f}" if claim.total_charges else "",
            ),
            ("LOS", str(claim.los) if claim.los else ""),
            ("Principal Dx", claim.principal_dx.code if claim.principal_dx else ""),
            ("Line Count", str(len(claim.lines)) if claim.lines else "0"),
        ]

        for label, value in claim_info:
            ws.cell(row=current_row, column=1, value=label)
            ws.cell(row=current_row, column=2, value=value)
            ws.cell(row=current_row, column=1).border = THIN_BORDER
            ws.cell(row=current_row, column=2).border = THIN_BORDER
            current_row += 1

        current_row += 1  # Add spacing

    # Error status
    if output.error:
        ws.cell(row=current_row, column=1, value="Error")
        ws.cell(row=current_row, column=2, value=output.error)
        ws.cell(row=current_row, column=1).font = Font(bold=True, color="FF0000")
        ws.cell(row=current_row, column=2).font = Font(color="FF0000")
        current_row += 2

    # Module status table
    ws.cell(row=current_row, column=1, value="Module")
    ws.cell(row=current_row, column=2, value="Status")
    ws.cell(row=current_row, column=3, value="Key Information")
    _style_header_row(ws, current_row, 3)
    current_row += 1

    # Define modules and their key info extractors
    modules = [
        ("IOCE", output.ioce, _get_ioce_summary),
        ("MCE", output.mce, _get_mce_summary),
        ("MS-DRG", output.msdrg, _get_msdrg_summary),
        ("HHAG", output.hhag, _get_hhag_summary),
        ("CMG/IRF Grouper", output.cmg, _get_cmg_summary),
        ("IPPS", output.ipps, _get_ipps_summary),
        ("OPPS", output.opps, _get_opps_summary),
        ("IPF/Psych", output.psych, _get_generic_summary),
        ("LTCH", output.ltch, _get_generic_summary),
        ("IRF Pricer", output.irf, _get_generic_summary),
        ("Hospice", output.hospice, _get_generic_summary),
        ("SNF", output.snf, _get_generic_summary),
        ("HHA", output.hha, _get_generic_summary),
        ("ESRD", output.esrd, _get_generic_summary),
        ("FQHC", output.fqhc, _get_fqhc_summary),
        ("IPSF", output.ipsf, _get_ipsf_summary),
        ("OPSF", output.opsf, _get_opsf_summary),
    ]

    for name, module_output, summary_func in modules:
        status = "✓ Processed" if module_output else "— Not Run"
        if name == "IPSF" or name == "OPSF":
            status = "✓ Used for Pricing" 
        key_info = summary_func(module_output) if module_output else ""

        ws.cell(row=current_row, column=1, value=name)
        ws.cell(row=current_row, column=2, value=status)
        ws.cell(row=current_row, column=3, value=key_info)

        for col in range(1, 4):
            ws.cell(row=current_row, column=col).border = THIN_BORDER

        if module_output:
            ws.cell(row=current_row, column=2).font = Font(color="008000")  # Green

        current_row += 1

    # Add generation timestamp
    current_row += 2
    ws.cell(
        row=current_row,
        column=1,
        value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    )
    ws.cell(row=current_row, column=1).font = Font(italic=True, color="808080")

    _auto_adjust_column_widths(ws)


def _get_ioce_summary(output: Any) -> str:
    """Get summary info for IOCE output."""
    if not output:
        return ""
    parts = []
    if hasattr(output, "processing_information") and output.processing_information:
        pi = output.processing_information
        if hasattr(pi, "return_code") and pi.return_code:
            parts.append(f"Return Code: {pi.return_code.code}")
        if hasattr(pi, "lines_processed"):
            parts.append(f"Lines: {pi.lines_processed}")
    return ", ".join(parts) if parts else "Processed"


def _get_mce_summary(output: Any) -> str:
    """Get summary info for MCE output."""
    if not output:
        return ""
    parts = []
    if hasattr(output, "return_code"):
        parts.append(f"Return Code: {output.return_code}")
    if hasattr(output, "diagnosis_codes") and output.diagnosis_codes:
        parts.append(f"Dx Codes: {len(output.diagnosis_codes)}")
    return ", ".join(parts) if parts else "Processed"


def _get_msdrg_summary(output: Any) -> str:
    """Get summary info for MS-DRG output."""
    if not output:
        return ""
    parts = []
    if hasattr(output, "final_drg_value") and output.final_drg_value:
        parts.append(f"DRG: {output.final_drg_value}")
    if hasattr(output, "final_mdc_value") and output.final_mdc_value:
        parts.append(f"MDC: {output.final_mdc_value}")
    return ", ".join(parts) if parts else "Processed"


def _get_hhag_summary(output: Any) -> str:
    """Get summary info for HHAG output."""
    if not output:
        return ""
    parts = []
    if hasattr(output, "hipps") and output.hipps:
        parts.append(f"HIPPS: {output.hipps}")
    return ", ".join(parts) if parts else "Processed"


def _get_cmg_summary(output: Any) -> str:
    """Get summary info for CMG/IRF Grouper output."""
    if not output:
        return ""
    parts = []
    if hasattr(output, "cmg") and output.cmg:
        parts.append(f"CMG: {output.cmg}")
    return ", ".join(parts) if parts else "Processed"


def _get_ipps_summary(output: Any) -> str:
    """Get summary info for IPPS output."""
    if not output:
        return ""
    parts = []
    if hasattr(output, "total_payment") and output.total_payment:
        parts.append(f"Total Payment: ${output.total_payment:,.2f}")
    return ", ".join(parts) if parts else "Processed"


def _get_opps_summary(output: Any) -> str:
    """Get summary info for OPPS output."""
    if not output:
        return ""
    parts = []
    if hasattr(output, "payment") and output.payment:
        parts.append(f"Total Payment: ${output.payment:,.2f}")
    if hasattr(output, "line_outputs") and output.line_outputs:
        parts.append(f"Lines: {len(output.line_outputs)}")
    return ", ".join(parts) if parts else "Processed"


def _get_fqhc_summary(output: Any) -> str:
    """Get summary info for FQHC output."""
    if not output:
        return ""
    parts = []
    if hasattr(output, "total_payment") and output.total_payment:
        parts.append(f"Total Payment: ${output.total_payment:,.2f}")
    if hasattr(output, "line_payment_data") and output.line_payment_data:
        line_count = len(output.line_payment_data)
        parts.append(f"Lines: {line_count}")
        # Sum up line-level payments
        line_payments = sum(line.payment or 0 for line in output.line_payment_data)
        if line_payments > 0:
            parts.append(f"Line Payments: ${line_payments:,.2f}")
    if hasattr(output, "coinsurance_amount") and output.coinsurance_amount:
        parts.append(f"Coinsurance: ${output.coinsurance_amount:,.2f}")
    return ", ".join(parts) if parts else "Processed"


def _get_ipsf_summary(output: Any) -> str:
    """Get summary info for IPSF output."""
    if not output:
        return ""
    parts = []
    if hasattr(output, "provider_ccn") and output.provider_ccn:
        parts.append(f"CCN: {output.provider_ccn}")
    if hasattr(output, "provider_type") and output.provider_type:
        parts.append(f"Type: {output.provider_type}")
    return ", ".join(parts) if parts else "Used for Processing"


def _get_opsf_summary(output: Any) -> str:
    """Get summary info for OPSF output."""
    if not output:
        return ""
    parts = []
    if hasattr(output, "provider_ccn") and output.provider_ccn:
        parts.append(f"CCN: {output.provider_ccn}")
    if hasattr(output, "provider_type") and output.provider_type:
        parts.append(f"Type: {output.provider_type}")
    return ", ".join(parts) if parts else "Used for Processing"


def _get_generic_summary(output: Any) -> str:
    """Get generic summary for any output."""
    if not output:
        return ""
    # Try common payment fields
    for field in ["total_payment", "payment", "amount"]:
        if hasattr(output, field):
            value = getattr(output, field)
            if value is not None and isinstance(value, (int, float)):
                return f"Payment: ${value:,.2f}"
    return "Processed"


class ExcelExporter:
    """
    Export MyelinOutput to Excel format.

    This class provides methods to convert complex nested MyelinOutput
    structures into well-formatted Excel workbooks with multiple sheets.
    """

    def __init__(
        self,
        output: "MyelinOutput",
        claim: "Claim | None" = None,
    ):
        """
        Initialize the exporter with a MyelinOutput instance.

        Args:
            output: The MyelinOutput to export
            claim: Optional input Claim to include in the export
        """
        _ensure_openpyxl()
        self.output = output
        self.claim = claim

    def export(self, filepath: str | Path) -> None:
        """
        Export the output to an Excel file.

        Args:
            filepath: Path where the Excel file should be saved
        """
        wb = self._create_workbook()
        wb.save(filepath)

    def export_to_bytes(self) -> bytes:
        """
        Export the output to bytes (useful for web applications).

        Returns:
            Excel file content as bytes
        """
        wb = self._create_workbook()
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def _create_workbook(self) -> "openpyxl.Workbook":
        """Create the Excel workbook with all sheets."""
        wb = openpyxl.Workbook()

        # Create summary sheet first
        _create_summary_sheet(wb, self.output, self.claim)

        # Add claim input sheet if claim is provided
        if self.claim is not None:
            ws = wb.create_sheet(title="Claim Input")
            _write_model_to_sheet(ws, self.claim, "Input Claim")
            _auto_adjust_column_widths(ws)

        # Add sheets for each module that has data
        modules = [
            ("ioce", "IOCE", self.output.ioce),
            ("mce", "MCE", self.output.mce),
            ("msdrg", "MS-DRG", self.output.msdrg),
            ("hhag", "HHAG", self.output.hhag),
            ("cmg", "CMG", self.output.cmg),
            ("ipps", "IPPS", self.output.ipps),
            ("opps", "OPPS", self.output.opps),
            ("psych", "IPF-Psych", self.output.psych),
            ("ltch", "LTCH", self.output.ltch),
            ("irf", "IRF", self.output.irf),
            ("hospice", "Hospice", self.output.hospice),
            ("snf", "SNF", self.output.snf),
            ("hha", "HHA", self.output.hha),
            ("esrd", "ESRD", self.output.esrd),
            ("fqhc", "FQHC", self.output.fqhc),
            ("ipsf", "IPSF", self.output.ipsf),
            ("opsf", "OPSF", self.output.opsf),
        ]

        for attr_name, sheet_name, module_output in modules:
            if module_output is not None:
                ws = wb.create_sheet(title=sheet_name)
                title = f"{sheet_name} Output" if attr_name != "ipsf" and attr_name != "opsf" else f"{sheet_name} Used for Pricing"
                _write_model_to_sheet(ws, module_output, title)
                _auto_adjust_column_widths(ws)

        return wb


def export_to_excel(
    output: "MyelinOutput",
    filepath: str | Path,
    claim: "Claim | None" = None,
) -> None:
    """
    Convenience function to export MyelinOutput to Excel.

    Args:
        output: The MyelinOutput to export
        filepath: Path where the Excel file should be saved
        claim: Optional input Claim to include in the export

    Example:
        >>> from myelin.helpers.excel_exporter import export_to_excel
        >>> result = myelin.process(claim)
        >>> export_to_excel(result, "output.xlsx", claim=claim)
    """
    exporter = ExcelExporter(output, claim=claim)
    exporter.export(filepath)


def export_to_excel_bytes(
    output: "MyelinOutput",
    claim: "Claim | None" = None,
) -> bytes:
    """
    Convenience function to export MyelinOutput to Excel bytes.

    Useful for web applications that need to return the file as a response.

    Args:
        output: The MyelinOutput to export
        claim: Optional input Claim to include in the export

    Returns:
        Excel file content as bytes

    Example:
        >>> from myelin.helpers.excel_exporter import export_to_excel_bytes
        >>> result = myelin.process(claim)
        >>> excel_bytes = export_to_excel_bytes(result, claim=claim)
        >>> # Send excel_bytes as HTTP response
    """
    exporter = ExcelExporter(output, claim=claim)
    return exporter.export_to_bytes()

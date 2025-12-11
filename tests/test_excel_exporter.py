"""Tests for Excel exporter functionality."""

import pytest
from datetime import datetime
from pathlib import Path
from pydantic import BaseModel, Field

# Test helper models
from myelin.helpers.excel_exporter import (
    _flatten_model,
    _format_value,
    _humanize_key,
    _extract_list_items,
    _is_edit_list,
    _concatenate_edit_list,
    _is_simple_string_list,
    _concatenate_string_list,
)


class SampleNestedModel(BaseModel):
    """Test nested model."""
    value: str = "test"
    count: int = 42


class SampleListItemModel(BaseModel):
    """Test list item model."""
    name: str = ""
    amount: float = 0.0


class SampleEditItem(BaseModel):
    """Test edit item model (mimics IoceOutputEdit)."""
    edit: str = ""
    description: str = ""


class SampleModel(BaseModel):
    """Test model with various field types."""
    simple_field: str = "hello"
    number_field: int = 100
    float_field: float = 99.99
    bool_field: bool = True
    none_field: str | None = None
    nested: SampleNestedModel = Field(default_factory=SampleNestedModel)
    items: list[SampleListItemModel] = Field(default_factory=list)


class TestFormatValue:
    """Tests for _format_value function."""

    def test_none_value(self):
        assert _format_value(None) == ""

    def test_bool_true(self):
        assert _format_value(True) == "Yes"

    def test_bool_false(self):
        assert _format_value(False) == "No"

    def test_datetime(self):
        dt = datetime(2024, 1, 15, 10, 30)
        assert _format_value(dt) == dt

    def test_string(self):
        assert _format_value("test") == "test"

    def test_number(self):
        assert _format_value(42) == 42
        assert _format_value(3.14) == 3.14


class TestHumanizeKey:
    """Tests for _humanize_key function."""

    def test_simple_key(self):
        assert _humanize_key("simple") == "Simple"

    def test_snake_case(self):
        assert _humanize_key("my_field_name") == "My Field Name"

    def test_already_readable(self):
        assert _humanize_key("Test") == "Test"


class TestFlattenModel:
    """Tests for _flatten_model function."""

    def test_simple_fields(self):
        model = SampleModel()
        flat = _flatten_model(model)
        
        assert flat["simple_field"] == "hello"
        assert flat["number_field"] == 100
        assert flat["float_field"] == 99.99
        assert flat["bool_field"] == "Yes"
        assert flat["none_field"] == ""

    def test_nested_model(self):
        model = SampleModel()
        flat = _flatten_model(model)
        
        assert flat["nested_value"] == "test"
        assert flat["nested_count"] == 42

    def test_list_field_count(self):
        model = SampleModel(items=[
            SampleListItemModel(name="item1", amount=10.0),
            SampleListItemModel(name="item2", amount=20.0),
        ])
        flat = _flatten_model(model)
        
        assert flat["items_count"] == 2


class TestExtractListItems:
    """Tests for _extract_list_items function."""

    def test_empty_list(self):
        model = SampleModel()
        lists = _extract_list_items(model)
        
        # Empty list should not be extracted
        assert "items" not in lists

    def test_populated_list(self):
        model = SampleModel(items=[
            SampleListItemModel(name="item1", amount=10.0),
        ])
        lists = _extract_list_items(model)
        
        assert "items" in lists
        assert len(lists["items"]) == 1


class TestEditListHandling:
    """Tests for edit list concatenation functionality."""

    def test_is_edit_list_by_name(self):
        """Test detection of edit lists by field name."""
        items = [SampleEditItem(edit="E001")]
        assert _is_edit_list("edit_list", items) is True
        assert _is_edit_list("hcpcs_edit_list", items) is True
        assert _is_edit_list("claim_rejection_edit_list", items) is True

    def test_is_edit_list_by_attribute(self):
        """Test detection of edit lists by item attributes."""
        items = [SampleEditItem(edit="E001")]
        assert _is_edit_list("some_field", items) is True
        
        # Non-edit items should not be detected
        items = [SampleListItemModel(name="test")]
        assert _is_edit_list("some_field", items) is False

    def test_concatenate_edit_list_with_descriptions(self):
        """Test concatenation with both edit code and description."""
        items = [
            SampleEditItem(edit="E001", description="Invalid code"),
            SampleEditItem(edit="E002", description="Missing modifier"),
        ]
        result = _concatenate_edit_list(items)
        assert result == "E001: Invalid code\nE002: Missing modifier"

    def test_concatenate_edit_list_without_descriptions(self):
        """Test concatenation with only edit codes."""
        items = [
            SampleEditItem(edit="E001", description=""),
            SampleEditItem(edit="E002", description=""),
        ]
        result = _concatenate_edit_list(items)
        assert result == "E001\nE002"

    def test_concatenate_empty_list(self):
        """Test concatenation of empty list."""
        result = _concatenate_edit_list([])
        assert result == ""

    def test_flatten_model_with_edit_list(self):
        """Test that edit lists are concatenated in flattened output."""
        class ModelWithEditList(BaseModel):
            name: str = "test"
            edit_list: list[SampleEditItem] = Field(default_factory=list)
        
        model = ModelWithEditList(
            name="test",
            edit_list=[
                SampleEditItem(edit="E001", description="Error 1"),
                SampleEditItem(edit="E002", description="Error 2"),
            ]
        )
        flat = _flatten_model(model)
        
        # Edit list should be concatenated, not counted
        assert "edit_list" in flat
        assert flat["edit_list"] == "E001: Error 1\nE002: Error 2"
        assert "edit_list_count" not in flat

    def test_extract_list_items_excludes_edit_lists(self):
        """Test that edit lists are not extracted as separate tables."""
        class ModelWithEditList(BaseModel):
            name: str = "test"
            edit_list: list[SampleEditItem] = Field(default_factory=list)
            items: list[SampleListItemModel] = Field(default_factory=list)
        
        model = ModelWithEditList(
            edit_list=[SampleEditItem(edit="E001")],
            items=[SampleListItemModel(name="item1")],
        )
        lists = _extract_list_items(model)
        
        # Edit list should be excluded, regular list should be included
        assert "edit_list" not in lists
        assert "items" in lists


class TestStringListHandling:
    """Tests for simple string list concatenation (like modifiers)."""

    def test_is_simple_string_list_true(self):
        """Test detection of simple string lists."""
        assert _is_simple_string_list("modifiers", ["25", "59"]) is True
        assert _is_simple_string_list("cond_codes", ["C1", "C2"]) is True
        assert _is_simple_string_list("demo_codes", ["D1"]) is True

    def test_is_simple_string_list_false_for_empty(self):
        """Test empty lists are not detected."""
        assert _is_simple_string_list("modifiers", []) is False

    def test_is_simple_string_list_false_for_non_strings(self):
        """Test non-string lists are not detected."""
        assert _is_simple_string_list("items", [1, 2, 3]) is False
        assert _is_simple_string_list("objects", [SampleListItemModel()]) is False

    def test_concatenate_string_list(self):
        """Test concatenation of string lists."""
        result = _concatenate_string_list(["25", "59", "76"])
        assert result == "25; 59; 76"

    def test_concatenate_string_list_with_custom_delimiter(self):
        """Test concatenation with custom delimiter."""
        result = _concatenate_string_list(["A", "B", "C"], delimiter=", ")
        assert result == "A, B, C"

    def test_concatenate_string_list_empty(self):
        """Test concatenation of empty list."""
        assert _concatenate_string_list([]) == ""

    def test_concatenate_string_list_filters_empty_strings(self):
        """Test that empty strings are filtered out."""
        result = _concatenate_string_list(["25", "", "59", ""])
        assert result == "25; 59"

    def test_flatten_model_with_modifiers(self):
        """Test that modifier lists are concatenated in flattened output."""
        class LineItemModel(BaseModel):
            hcpcs: str = ""
            modifiers: list[str] = Field(default_factory=list)
        
        model = LineItemModel(
            hcpcs="99213",
            modifiers=["25", "59"],
        )
        flat = _flatten_model(model)
        
        # Modifiers should be concatenated, not counted
        assert "modifiers" in flat
        assert flat["modifiers"] == "25; 59"
        assert "modifiers_count" not in flat

    def test_flatten_model_with_cond_codes(self):
        """Test that condition codes are concatenated."""
        class ClaimModel(BaseModel):
            cond_codes: list[str] = Field(default_factory=list)
        
        model = ClaimModel(cond_codes=["C1", "C2", "C3"])
        flat = _flatten_model(model)
        
        assert "cond_codes" in flat
        assert flat["cond_codes"] == "C1; C2; C3"

    def test_extract_list_items_excludes_string_lists(self):
        """Test that string lists are not extracted as separate tables."""
        class ModelWithStringList(BaseModel):
            name: str = "test"
            modifiers: list[str] = Field(default_factory=list)
            items: list[SampleListItemModel] = Field(default_factory=list)
        
        model = ModelWithStringList(
            modifiers=["25", "59"],
            items=[SampleListItemModel(name="item1")],
        )
        lists = _extract_list_items(model)
        
        # String list should be excluded, BaseModel list should be included
        assert "modifiers" not in lists
        assert "items" in lists


class TestExcelExporter:
    """Integration tests for ExcelExporter."""

    @pytest.fixture
    def sample_output(self):
        """Create a sample MyelinOutput for testing."""
        from myelin.core import MyelinOutput
        from myelin.msdrg.msdrg_output import MsdrgOutput
        
        return MyelinOutput(
            msdrg=MsdrgOutput(
                claim_id="test-claim-001",
                drg_version="42.0",
                final_drg_value="470",
                final_drg_description="Major Hip and Knee Joint Replacement",
                final_mdc_value="08",
                final_mdc_description="Diseases and Disorders of the Musculoskeletal System",
            )
        )

    @pytest.mark.skipif(
        not pytest.importorskip("openpyxl", reason="openpyxl not installed"),
        reason="openpyxl required for this test"
    )
    def test_export_to_bytes(self, sample_output):
        """Test exporting to bytes."""
        try:
            from myelin.helpers.excel_exporter import ExcelExporter
            
            exporter = ExcelExporter(sample_output)
            excel_bytes = exporter.export_to_bytes()
            
            assert isinstance(excel_bytes, bytes)
            assert len(excel_bytes) > 0
            # Check for Excel file signature (ZIP format)
            assert excel_bytes[:4] == b'PK\x03\x04'
        except ImportError:
            pytest.skip("openpyxl not installed")

    @pytest.mark.skipif(
        not pytest.importorskip("openpyxl", reason="openpyxl not installed"),
        reason="openpyxl required for this test"
    )
    def test_export_to_file(self, sample_output, tmp_path):
        """Test exporting to a file."""
        try:
            from myelin.helpers.excel_exporter import export_to_excel
            
            filepath = tmp_path / "test_output.xlsx"
            export_to_excel(sample_output, filepath)
            
            assert filepath.exists()
            assert filepath.stat().st_size > 0
        except ImportError:
            pytest.skip("openpyxl not installed")

    @pytest.mark.skipif(
        not pytest.importorskip("openpyxl", reason="openpyxl not installed"),
        reason="openpyxl required for this test"
    )
    def test_myelin_output_to_excel_method(self, sample_output, tmp_path):
        """Test the to_excel method on MyelinOutput."""
        try:
            filepath = tmp_path / "test_method.xlsx"
            sample_output.to_excel(str(filepath))
            
            assert filepath.exists()
        except ImportError:
            pytest.skip("openpyxl not installed")

    @pytest.fixture
    def sample_claim(self):
        """Create a sample Claim for testing."""
        from datetime import datetime
        from myelin.input.claim import Claim, DiagnosisCode, LineItem
        
        return Claim(
            claimid="TEST-CLAIM-001",
            from_date=datetime(2024, 1, 1),
            thru_date=datetime(2024, 1, 5),
            bill_type="111",
            total_charges=5000.00,
            los=4,
            principal_dx=DiagnosisCode(code="M17.11"),
            lines=[
                LineItem(
                    revenue_code="0360",
                    hcpcs="27447",
                    charges=4500.00,
                ),
                LineItem(
                    revenue_code="0250",
                    hcpcs="",
                    charges=500.00,
                ),
            ],
        )

    @pytest.mark.skipif(
        not pytest.importorskip("openpyxl", reason="openpyxl not installed"),
        reason="openpyxl required for this test"
    )
    def test_export_with_claim_input(self, sample_output, sample_claim, tmp_path):
        """Test exporting with claim input included."""
        try:
            import openpyxl
            from myelin.helpers.excel_exporter import export_to_excel
            
            filepath = tmp_path / "test_with_claim.xlsx"
            export_to_excel(sample_output, filepath, claim=sample_claim)
            
            assert filepath.exists()
            
            # Verify the workbook has a Claim Input sheet
            wb = openpyxl.load_workbook(filepath)
            assert "Claim Input" in wb.sheetnames
            assert "Summary" in wb.sheetnames
            
            # Check summary sheet has claim info
            summary_ws = wb["Summary"]
            # Look for claim ID in the summary
            found_claim_id = False
            for row in summary_ws.iter_rows(values_only=True):
                if row and "TEST-CLAIM-001" in str(row):
                    found_claim_id = True
                    break
            assert found_claim_id, "Claim ID not found in summary sheet"
            
        except ImportError:
            pytest.skip("openpyxl not installed")

    @pytest.mark.skipif(
        not pytest.importorskip("openpyxl", reason="openpyxl not installed"),
        reason="openpyxl required for this test"
    )
    def test_export_with_claim_to_bytes(self, sample_output, sample_claim):
        """Test exporting with claim input to bytes."""
        try:
            from myelin.helpers.excel_exporter import ExcelExporter
            
            exporter = ExcelExporter(sample_output, claim=sample_claim)
            excel_bytes = exporter.export_to_bytes()
            
            assert isinstance(excel_bytes, bytes)
            assert len(excel_bytes) > 0
        except ImportError:
            pytest.skip("openpyxl not installed")

    @pytest.mark.skipif(
        not pytest.importorskip("openpyxl", reason="openpyxl not installed"),
        reason="openpyxl required for this test"
    )
    def test_to_excel_method_with_claim(self, sample_output, sample_claim, tmp_path):
        """Test the to_excel method with claim parameter."""
        try:
            filepath = tmp_path / "test_method_with_claim.xlsx"
            sample_output.to_excel(str(filepath), claim=sample_claim)
            
            assert filepath.exists()
        except ImportError:
            pytest.skip("openpyxl not installed")

    def test_import_error_without_openpyxl(self):
        """Test that proper error is raised when openpyxl is not available."""
        from myelin.helpers import excel_exporter
        
        # Save original state
        original_available = excel_exporter.OPENPYXL_AVAILABLE
        
        try:
            # Simulate openpyxl not being available
            excel_exporter.OPENPYXL_AVAILABLE = False
            
            with pytest.raises(ImportError) as exc_info:
                excel_exporter._ensure_openpyxl()
            
            assert "openpyxl" in str(exc_info.value)
            assert "pip install" in str(exc_info.value)
        finally:
            # Restore original state
            excel_exporter.OPENPYXL_AVAILABLE = original_available
